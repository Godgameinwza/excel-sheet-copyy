import os
import customtkinter as ctk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook

# --- ฟังก์ชัน ---
def load_template_values():
    """โหลดค่าจาก Sheet ต้นแบบมาแสดงใน input"""
    file_path = file_path_entry.get()
    base_sheet = base_entry.get()
    
    if not file_path or not base_sheet:
        messagebox.showwarning("ข้อมูลไม่ครบ", "กรุณาเลือกไฟล์และระบุชื่อ Sheet ต้นแบบก่อน")
        return
    
    if not os.path.exists(file_path):
        messagebox.showerror("ไฟล์ไม่พบ", "ไม่พบไฟล์ Excel ที่เลือก")
        return
    
    try:
        wb = load_workbook(file_path, data_only=True)
        if base_sheet not in wb.sheetnames:
            messagebox.showerror("ไม่พบ Sheet", f"ไม่พบ sheet '{base_sheet}' ในไฟล์นี้")
            wb.close()
            return
        
        ws = wb[base_sheet]
        
        # อ่านค่าจาก cell I3 และ J11
        num_val = ws["I3"].value
        date_val = ws["J11"].value
        month_val = ws["B18"].value
        
        # แสดงค่าใน input
        num_entry.delete(0, 'end')
        if num_val:
            num_entry.insert(0, str(num_val))
        
        date_entry.delete(0, 'end')
        if date_val:
            date_entry.insert(0, str(date_val))

        month_entry.delete(0, 'end')
        if month_val:
            month_entry.insert(0, str(month_val))
        
        wb.close()
        
    except Exception as e:
        messagebox.showerror("เกิดข้อผิดพลาด", f"ไม่สามารถโหลดค่าได้:\n{str(e)}")

def select_file():
    file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file:
        file_path_entry.delete(0, 'end')
        file_path_entry.insert(0, file)

def validate_sheet_name(name):
    """ตรวจสอบชื่อ Sheet ว่าถูกต้องตามกฎของ Excel"""
    if not name or len(name) > 31:
        return False, "ชื่อ Sheet ต้องมีความยาว 1-31 ตัวอักษร"
    
    invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
    for char in invalid_chars:
        if char in name:
            return False, f"ชื่อ Sheet ไม่สามารถมีอักขระพิเศษ: \\ / * ? : [ ]"
    
    return True, ""

def clear_inputs():
    """ล้างค่าในช่อง input หลังบันทึกสำเร็จ"""
    new_entry.delete(0, 'end')
    num_entry.delete(0, 'end')
    date_entry.delete(0, 'end')
    month_entry.delete(0, 'end')

def add_sheet_to_file():
    file_path = file_path_entry.get()
    base_sheet = base_entry.get()
    new_sheet = new_entry.get().strip()
    num_value = num_entry.get()
    date_value = date_entry.get()
    month_value = month_entry.get()

    # ตรวจสอบข้อมูลพื้นฐาน
    if not file_path or not base_sheet or not new_sheet:
        messagebox.showwarning("ข้อมูลไม่ครบ", "กรุณากรอกข้อมูลให้ครบ")
        return

    # ตรวจสอบไฟล์
    if not os.path.exists(file_path):
        messagebox.showerror("ไฟล์ไม่พบ", "ไม่พบไฟล์ Excel ที่เลือก")
        return

    # ตรวจสอบชื่อ Sheet ใหม่
    is_valid, error_msg = validate_sheet_name(new_sheet)
    if not is_valid:
        messagebox.showerror("ชื่อ Sheet ไม่ถูกต้อง", error_msg)
        return

    try:
        wb = load_workbook(file_path)
        
        # ตรวจสอบว่ามี Sheet ต้นแบบหรือไม่
        if base_sheet not in wb.sheetnames:
            messagebox.showerror("ไม่พบ Sheet", f"ไม่พบ sheet '{base_sheet}' ในไฟล์นี้")
            return

        # ตรวจสอบว่า Sheet ใหม่ซ้ำหรือไม่
        if new_sheet in wb.sheetnames:
            result = messagebox.askyesno("Sheet ซ้ำ", f"มี Sheet '{new_sheet}' อยู่แล้ว\nต้องการแทนที่หรือไม่?")
            if result:
                del wb[new_sheet]
            else:
                return

        # คัดลอก Sheet
        template = wb[base_sheet]
        new_ws = wb.copy_worksheet(template)
        new_ws.title = new_sheet
        
        # ใส่ค่าในเซลล์
        if num_value:
            new_ws["I3"] = num_value
        if date_value:
            new_ws["J11"] = date_value
        if month_value:
            new_ws["B18"] = month_value

        # บันทึกไฟล์
        wb.save(file_path)
        wb.close()
        
        messagebox.showinfo("เสร็จสิ้น", f"เพิ่ม Sheet '{new_sheet}' สำเร็จ")
        clear_inputs()
        
    except PermissionError:
        messagebox.showerror("ไม่สามารถบันทึกได้", "ไฟล์กำลังเปิดอยู่ กรุณาปิดไฟล์แล้วลองใหม่")
    except Exception as e:
        messagebox.showerror("เกิดข้อผิดพลาด", str(e))

# --- UI ---
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

root = ctk.CTk()
root.title("📄 Excel Sheet Copier")
root.geometry("550x650")
root.resizable(False, False)

frame = ctk.CTkFrame(root, corner_radius=15)
frame.pack(padx=20, pady=20, fill="both", expand=True)

# เลือกไฟล์
ctk.CTkLabel(frame, text="📁 เลือกไฟล์ Excel", font=("Arial", 14, "bold")).pack(pady=(10, 5))
file_path_entry = ctk.CTkEntry(frame, width=400)
file_path_entry.pack(pady=5)
ctk.CTkButton(frame, text="เลือกไฟล์", command=select_file).pack(pady=5)

# Sheet ต้นแบบ + ปุ่มโหลด
ctk.CTkLabel(frame, text="📑 ชื่อ Sheet ต้นแบบ", font=("Arial", 12)).pack(pady=(10, 5))
base_frame = ctk.CTkFrame(frame, fg_color="transparent")
base_frame.pack(pady=5)
base_entry = ctk.CTkEntry(base_frame, width=240)
base_entry.pack(side="left", padx=(0, 5))
ctk.CTkButton(base_frame, text="🔄 โหลดค่าเดิม", command=load_template_values, width=140, fg_color="#2196F3").pack(side="left")

# Sheet ใหม่
ctk.CTkLabel(frame, text="➕ ชื่อ Sheet ใหม่", font=("Arial", 12)).pack(pady=(10, 5))
new_entry = ctk.CTkEntry(frame, width=300)
new_entry.pack(pady=5)

# เลขที่ I3
ctk.CTkLabel(frame, text="🔢 เลขที่ (I3)", font=("Arial", 12)).pack(pady=(10, 5))
num_entry = ctk.CTkEntry(frame, width=300, placeholder_text="กรอกเลขที่หรือกด 'โหลดค่าเดิม' เพื่อดูค่าปัจจุบัน")
num_entry.pack(pady=5)

# Date J11
ctk.CTkLabel(frame, text="📅 วันที่ (J11)", font=("Arial", 12)).pack(pady=(10, 5))
date_entry = ctk.CTkEntry(frame, width=300, placeholder_text="กรอกวันที่หรือกด 'โหลดค่าเดิม' เพื่อดูค่าปัจจุบัน")
date_entry.pack(pady=5)

ctk.CTkLabel(frame, text="📅 เดือน (B18)", font=("Arial", 12)).pack(pady=(10, 5))
month_entry = ctk.CTkEntry(frame, width=300, placeholder_text="กรอกวันที่หรือกด 'โหลดค่าเดิม' เพื่อดูค่าปัจจุบัน")
month_entry.pack(pady=5)

# ปุ่มเพิ่ม Sheet
ctk.CTkButton(frame, text="✨ เพิ่ม Sheet", command=add_sheet_to_file, fg_color="#4CAF50", height=40, font=("Arial", 13, "bold")).pack(pady=20)

root.mainloop()